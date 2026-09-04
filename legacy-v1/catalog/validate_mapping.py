#!/usr/bin/env python3
"""
Validate the mapping against the DDL — the fifth checker.

Four times today one vocabulary was written by hand in two files and drifted in
silence. mapping.py is the fifth opportunity: it names sheet headers on one side
and Postgres columns on the other, and nothing stops either from moving.

This proves three things without a database, a sheet, or an API key:

  1. every db column in the mapping exists in 001_schema.sql
  2. every mapped CHECK-constrained column allows exactly the values the
     mapping's one_of() permits — no more, no less
  3. every NOT NULL column without a default is covered by a required mapping

Run it in CI alongside compose.py --check and validate_scenarios.py.

Usage:
    python catalog/validate_mapping.py
    python catalog/validate_mapping.py --sheet path/to/urace-catalogo-template.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

from mapping import SPECS  # noqa: E402

DDL = ROOT / "db" / "001_schema.sql"


def _split_columns(body: str) -> list[str]:
    """Split a CREATE TABLE body into column definitions.

    Splitting on newlines misses multi-line definitions, and a CHECK whose
    values sit on the next line then parses as having no constraint at all —
    which makes the checker silently stop checking. Found by negative test:
    misaligning an enum passed clean. Split on top-level commas instead.
    """
    parts, depth, current = [], 0, []
    for char in body:
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        if char == "," and depth == 0:
            parts.append("".join(current))
            current = []
        else:
            current.append(char)
    parts.append("".join(current))
    return [re.sub(r"--.*", "", p).strip() for p in parts if p.strip()]


def parse_ddl(text: str) -> dict:
    """Column name -> {check_values, not_null, has_default} per table."""
    tables: dict[str, dict] = {}
    for match in re.finditer(
            r"CREATE TABLE (\w+)\s*\((.*?)\n\);", text, re.DOTALL):
        table, body = match.group(1), match.group(2)
        cols: dict[str, dict] = {}
        for definition in _split_columns(body):
            flat = " ".join(definition.split())
            if not flat or flat.startswith(("CONSTRAINT", "UNIQUE", "PRIMARY", "CHECK")):
                continue
            name = flat.split()[0]
            if not re.match(r"^[a-z_]+$", name):
                continue
            allowed = None
            inline = re.search(rf"CHECK\s*\(\s*{name}\s+IN\s*\((.*?)\)", flat)
            if inline:
                allowed = set(re.findall(r"'([^']+)'", inline.group(1)))
            cols[name] = {
                "check": allowed,
                "not_null": "NOT NULL" in flat,
                "default": "DEFAULT" in flat,
            }
        tables[table] = cols
    return tables


def one_of_values(coerce) -> set | None:
    """Recover the allowed set from a one_of() closure."""
    cells = getattr(coerce, "__closure__", None)
    if not cells:
        return None
    for cell in cells:
        if isinstance(cell.cell_contents, tuple) and all(
                isinstance(v, str) for v in cell.cell_contents):
            return set(cell.cell_contents)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", type=Path, help="Also check headers against the template.")
    args = ap.parse_args()

    if not DDL.exists():
        sys.exit(f"ERROR: {DDL} not found")
    tables = parse_ddl(DDL.read_text(encoding="utf-8"))
    problems: list[str] = []

    for spec in SPECS:
        cols = tables.get(spec.table)
        if cols is None:
            problems.append(f"{spec.tab}: table '{spec.table}' is not in the DDL")
            continue

        mapped = set()
        for col in spec.columns:
            mapped.add(col.db)

            # 1 — the column exists
            if col.db not in cols:
                problems.append(
                    f"{spec.tab}: maps '{col.sheet}' to {spec.table}.{col.db}, "
                    f"which does not exist")
                continue

            # 2 — the vocabularies agree, in both directions
            ddl_values = cols[col.db]["check"]
            map_values = one_of_values(col.coerce)
            if ddl_values and map_values and ddl_values != map_values:
                only_ddl = sorted(ddl_values - map_values)
                only_map = sorted(map_values - ddl_values)
                detail = []
                if only_ddl:
                    detail.append(f"DDL allows {only_ddl} the mapping rejects")
                if only_map:
                    detail.append(f"mapping allows {only_map} the DDL rejects")
                problems.append(f"{spec.tab}.{col.db}: " + "; ".join(detail))
            elif ddl_values and not map_values:
                problems.append(
                    f"{spec.tab}.{col.db}: DDL constrains this to "
                    f"{sorted(ddl_values)} but the mapping accepts free text")

        # 3 — nothing mandatory is left unmapped
        for name, meta in cols.items():
            if name in mapped or meta["default"] or not meta["not_null"]:
                continue
            if name in ("id", "created_at", "updated_at", "synced_at",
                        "source_sheet_row", "source_sheet_tab",
                        "program_id", "segment_id"):
                continue
            problems.append(
                f"{spec.tab}: {spec.table}.{name} is NOT NULL with no default "
                f"and nothing in the sheet fills it")

    # optional — headers against the template workbook
    if args.sheet:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.sheet, read_only=True)
            for spec in SPECS:
                if spec.tab not in wb.sheetnames:
                    problems.append(f"template is missing tab '{spec.tab}'")
                    continue
                headers = {c.value for c in next(wb[spec.tab].iter_rows(max_row=1))}
                for col in spec.columns:
                    if col.sheet not in headers:
                        problems.append(
                            f"{spec.tab}: mapping expects column '{col.sheet}', "
                            f"which is not in the template")
        except ImportError:
            print("note: openpyxl not installed, skipping the template check\n")

    if problems:
        print(f"FAILED — {len(problems)} mapping problems:\n")
        for p in problems:
            print(f"  {p}")
        print("\nThe sheet, mapping.py and 001_schema.sql describe one catalog.")
        print("When they disagree, the sync writes something nobody intended.")
        return 1

    total = sum(len(s.columns) for s in SPECS)
    print("OK — mapping agrees with the DDL.")
    print(f"  {len(SPECS)} tabs, {total} mapped columns")
    for spec in SPECS:
        req = sum(1 for c in spec.columns if c.required)
        enums = sum(1 for c in spec.columns if one_of_values(c.coerce))
        print(f"  {spec.tab:12} -> {spec.table:16} "
              f"{len(spec.columns):2} columns, {req} required, {enums} enum-checked")
    return 0


if __name__ == "__main__":
    sys.exit(main())
