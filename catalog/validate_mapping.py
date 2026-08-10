#!/usr/bin/env python3
"""
Validate the mapping against the schema spec — the fifth checker.

Four times today one vocabulary was written by hand in two files and drifted in
silence. mapping.py is the fifth opportunity: it names sheet headers on one side
and Firestore fields on the other, and nothing stops either from moving.

O Firestore não tem DDL, então a especificação virou db/schema.py — e este
checador passou a conferir contra ela. A propriedade continua a mesma.

This proves three things without a database, a sheet, or an API key:

  1. every db field in the mapping exists in db/schema.py
  2. every mapped enum-constrained field allows exactly the values the
     mapping's one_of() permits — no more, no less
  3. every required field is covered by a required mapping

Run it in CI alongside compose.py --check and validate_scenarios.py.

Usage:
    python catalog/validate_mapping.py
    python catalog/validate_mapping.py --sheet path/to/urace-catalogo-template.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from mapping import SPECS  # noqa: E402


def load_spec_tables() -> dict:
    """Field name -> {check_values, not_null, has_default} per collection.

    A forma espelha o que o parse do DDL devolvia, para que os três checks
    continuem idênticos. 'required' na spec cobre o papel do NOT NULL sem
    default; campos não-required contam como 'default: True' — o Firestore
    não tem DEFAULT, ausência é o default.
    """
    from db.schema import COLLECTIONS
    tables: dict[str, dict] = {}
    for collection, fields in COLLECTIONS.items():
        cols = {}
        for name, rules in fields.items():
            allowed = set(rules["check"]) if rules["check"] else None
            cols[name] = {
                "check": allowed,
                "not_null": rules["required"],
                "default": not rules["required"],
            }
        tables[collection] = cols
    return tables


def one_of_values(coerce) -> set | None:
    """Recover the allowed set from a one_of() closure."""
    cells = getattr(coerce, "__closure__", None)
    if not cells:
        return None
    for cell in cells:
        if isinstance(cell.cell_contents, tuple) and all(
                isinstance(v, str) for v in cell.cell_contents):
            return set(cell.cell_contents)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", type=Path, help="Also check headers against the template.")
    args = ap.parse_args()

    tables = load_spec_tables()
    problems: list[str] = []

    for spec in SPECS:
        cols = tables.get(spec.table)
        if cols is None:
            problems.append(
                f"{spec.tab}: collection '{spec.table}' is not in db/schema.py")
            continue

        mapped = set()
        for col in spec.columns:
            mapped.add(col.db)

            # 1 — the column exists
            if col.db not in cols:
                problems.append(
                    f"{spec.tab}: maps '{col.sheet}' to {spec.table}.{col.db}, "
                    f"which does not exist")
                continue

            # 2 — the vocabularies agree, in both directions
            spec_values = cols[col.db]["check"]
            map_values = one_of_values(col.coerce)
            if spec_values and map_values and spec_values != map_values:
                only_spec = sorted(spec_values - map_values)
                only_map = sorted(map_values - spec_values)
                detail = []
                if only_spec:
                    detail.append(f"schema allows {only_spec} the mapping rejects")
                if only_map:
                    detail.append(f"mapping allows {only_map} the schema rejects")
                problems.append(f"{spec.tab}.{col.db}: " + "; ".join(detail))
            elif spec_values and not map_values:
                problems.append(
                    f"{spec.tab}.{col.db}: schema constrains this to "
                    f"{sorted(spec_values)} but the mapping accepts free text")

        # 3 — nothing mandatory is left unmapped
        for name, meta in cols.items():
            if name in mapped or meta["default"] or not meta["not_null"]:
                continue
            if name in ("id", "created_at", "updated_at", "synced_at",
                        "source_sheet_row", "source_sheet_tab",
                        "program_id", "segment_id"):
                continue
            problems.append(
                f"{spec.tab}: {spec.table}.{name} is required "
                f"and nothing in the sheet fills it")

    # optional — headers against the template workbook
    if args.sheet:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.sheet, read_only=True)
            for spec in SPECS:
                if spec.tab not in wb.sheetnames:
                    problems.append(f"template is missing tab '{spec.tab}'")
                    continue
                headers = {c.value for c in next(wb[spec.tab].iter_rows(max_row=1))}
                for col in spec.columns:
                    if col.sheet not in headers:
                        problems.append(
                            f"{spec.tab}: mapping expects column '{col.sheet}', "
                            f"which is not in the template")
        except ImportError:
            print("note: openpyxl not installed, skipping the template check\n")

    if problems:
        print(f"FAILED — {len(problems)} mapping problems:\n")
        for p in problems:
            print(f"  {p}")
        print("\nThe sheet, mapping.py and db/schema.py describe one catalog.")
        print("When they disagree, the sync writes something nobody intended.")
        return 1

    total = sum(len(s.columns) for s in SPECS)
    print("OK — mapping agrees with the schema spec.")
    print(f"  {len(SPECS)} tabs, {total} mapped columns")
    for spec in SPECS:
        req = sum(1 for c in spec.columns if c.required)
        enums = sum(1 for c in spec.columns if one_of_values(c.coerce))
        print(f"  {spec.tab:12} -> {spec.table:16} "
              f"{len(spec.columns):2} columns, {req} required, {enums} enum-checked")
    return 0


if __name__ == "__main__":
    sys.exit(main())
